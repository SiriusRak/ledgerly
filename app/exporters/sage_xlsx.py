"""Sage-compatible XLSX export — partie double, 13 columns."""

from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font

from app.db import get_supabase

HEADERS = [
    "Compte", "Date", "Journal", "N\u00b0Piece", "R\u00e9f\u00e9rence",
    "Tiers", "Libell\u00e9s", "Lettrage", "D\u00e9bit", "Cr\u00e9dit",
    "Solde", "Mois", "Observation",
]


def _fmt_date(d: str | None) -> str:
    """Convert YYYY-MM-DD to DD/MM/YYYY."""
    if not d:
        return ""
    try:
        return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return d


def _fmt_mois(d: str | None) -> str:
    if not d:
        return ""
    try:
        return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%m/%Y")
    except ValueError:
        return d


def build_xlsx(month: str, dossier_client_id: str | None = None) -> bytes:
    """Build Sage 13-col XLSX for a given month (YYYY-MM), optionally filtered by dossier client."""
    sb = get_supabase()
    start = f"{month}-01"
    y, m = int(month[:4]), int(month[5:7])
    end = f"{y + 1}-01-01" if m == 12 else f"{y}-{m + 1:02d}-01"

    q = (
        sb.table("invoices")
        .select("*, suppliers(name)")
        .eq("state", "done")
        .gte("invoice_date", start)
        .lt("invoice_date", end)
        .order("invoice_date")
    )
    if dossier_client_id:
        q = q.eq("dossier_client_id", dossier_client_id)

    rows = q.execute().data or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ecritures"

    # Header row
    bold = Font(bold=True)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    r = 2
    for inv in rows:
        supplier_name = ""
        if inv.get("suppliers") and isinstance(inv["suppliers"], dict):
            supplier_name = inv["suppliers"].get("name", "")
        if not supplier_name:
            supplier_name = inv.get("supplier_name_raw", "")

        date_str = _fmt_date(inv.get("invoice_date"))
        mois_str = _fmt_mois(inv.get("invoice_date"))
        piece = inv.get("invoice_number", "")
        libelle = f"{supplier_name} - {piece}" if piece else supplier_name
        compte_charge = inv.get("compte", "606")
        amount_ht = float(inv.get("amount_ht") or 0)
        amount_tva = float(inv.get("amount_tva") or 0)
        amount_ttc = float(inv.get("amount_ttc") or 0)

        # Line 1 — charge
        ws.cell(r, 1, value=compte_charge)
        ws.cell(r, 2, value=date_str)
        ws.cell(r, 3, value="HA")
        ws.cell(r, 4, value=piece)
        ws.cell(r, 6, value=supplier_name)
        ws.cell(r, 7, value=libelle)
        ws.cell(r, 9, value=round(amount_ht, 2))
        ws.cell(r, 12, value=mois_str)
        r += 1

        # Line 2 — TVA (only if > 0)
        if amount_tva > 0:
            ws.cell(r, 1, value="44566")
            ws.cell(r, 2, value=date_str)
            ws.cell(r, 3, value="HA")
            ws.cell(r, 4, value=piece)
            ws.cell(r, 6, value=supplier_name)
            ws.cell(r, 7, value=libelle)
            ws.cell(r, 9, value=round(amount_tva, 2))
            ws.cell(r, 12, value=mois_str)
            r += 1

        # Line 3 — fournisseur (credit)
        ws.cell(r, 1, value="401")
        ws.cell(r, 2, value=date_str)
        ws.cell(r, 3, value="HA")
        ws.cell(r, 4, value=piece)
        ws.cell(r, 6, value=supplier_name)
        ws.cell(r, 7, value=libelle)
        ws.cell(r, 10, value=round(amount_ttc, 2))
        ws.cell(r, 12, value=mois_str)
        r += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
